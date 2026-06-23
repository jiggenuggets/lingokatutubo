"""
Translation dataset loader for LingoKatutubo phrasebook / translation memory.
Supports cross-lingual lookup between:
  English, Filipino, Cebuano, and Bagobo-Tagabawa.

This dataset is intentionally used as a phrasebook first. It is not large
enough to claim high-accuracy neural translation by itself.

Loads from (in priority order):
  1. translation_data.json  (utf-8-sig to handle BOM)
  2. Any .csv file in the same directory
  3. Any .xlsx file in the same directory
"""

import csv as _csv
import glob as _glob
import json
import os
import re
import unicodedata
from typing import Dict, List, Optional

from .display_utils import clean_invisible_unicode, safe_print

SUPPORTED_LANGS = ["english", "tagabawa", "filipino", "cebuano"]
UNKNOWN_FOR_REVIEW = "[UNKNOWN_FOR_REVIEW]"

_LANG_ALIASES = {
    "bagobo": "tagabawa",
    "tagabawa": "tagabawa",
    "bagobo-tagabawa": "tagabawa",
    "bagobo_tagabawa": "tagabawa",
    "bagobo tagabawa": "tagabawa",
    "bgs": "tagabawa",
    "english": "english",
    "en": "english",
    "eng": "english",
    "filipino": "filipino",
    "tagalog": "filipino",
    "fil": "filipino",
    "tgl": "filipino",
    "cebuano": "cebuano",
    "ceb": "cebuano",
    "bisaya": "cebuano",
    "visayan": "cebuano",
}

# --- Column normalization for CSV/Excel loaders ---
_COL_ALIASES = {
    "tagalog": "filipino_source",
    "bagobo": "tagabawa_source",
    "bagobo-tagabawa": "tagabawa_source",
    "bagobo_tagabawa": "tagabawa_source",
    "bagobo tagabawa": "tagabawa_source",
    "bgs": "tagabawa_source",
    "en": "english_source",
    "eng": "english_source",
    "fil": "filipino_source",
    "tgl": "filipino_source",
    "ceb": "cebuano_source",
    "bisaya": "cebuano_source",
    "visayan": "cebuano_source",
}


def _normalize_column(header: str) -> Optional[str]:
    """Map a raw CSV/Excel column header to its canonical {lang}_source key."""
    raw = str(header or "").strip().lower()
    if raw in (f"{l}_source" for l in SUPPORTED_LANGS):
        return raw

    h = _normalize_alias_key(header)
    if h in _COL_ALIASES:
        return _COL_ALIASES[h]
    canonical_lang = _normalize_lang(h)
    if canonical_lang in SUPPORTED_LANGS:
        return f"{canonical_lang}_source"
    return None


def _normalize_alias_key(value: str) -> str:
    return re.sub(r"\s+", " ", clean_invisible_unicode(value).strip().lower().replace("_", " "))


def _strip_diacritics(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def _normalize_lookup_text(value: str, *, strip_diacritics: bool = False) -> str:
    text = clean_invisible_unicode(value).strip().lower()
    if strip_diacritics:
        text = _strip_diacritics(text)
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _lookup_variants(value: str) -> List[str]:
    raw = clean_invisible_unicode(value).strip().lower()
    variants = [
        raw,
        _normalize_lookup_text(value),
        _normalize_lookup_text(value, strip_diacritics=True),
    ]
    output: List[str] = []
    for variant in variants:
        if variant and variant not in output:
            output.append(variant)
    return output


class TranslationDataset:
    """Manages cross-lingual phrase lookups from the multilingual phrasebook."""

    def __init__(self, dataset_path: Optional[str] = None):
        if dataset_path is None:
            base = os.path.dirname(__file__)
            candidates = [
                os.path.join(base, "translation_data.json"),
                os.path.join(base, "data", "translation_dataset.json"),
                os.path.join(base, "data", "translation_data.json"),
            ]
            dataset_path = next(
                (p for p in candidates if os.path.exists(p)), candidates[0]
            )

        self.dataset_path = dataset_path
        self.data: List[Dict] = []
        self.metadata: Dict = {}

        # Per-language indices: lang -> normalized_text -> [rows]
        self._phrase_indices: Dict[str, Dict[str, List[Dict]]] = {}
        # Per-language word indices: lang -> word -> [rows]
        self._word_indices: Dict[str, Dict[str, List[Dict]]] = {}

        self.is_loaded = False

        self.load()

    # ------------------------------------------------------------------
    # Loading
    # ------------------------------------------------------------------

    def load(self):
        # --- Attempt 1: JSON (the configured path) ---
        json_loaded = False
        if os.path.exists(self.dataset_path):
            safe_print(f"[Translation] Loading dataset from: {self.dataset_path}")
            try:
                with open(self.dataset_path, "r", encoding="utf-8-sig") as f:
                    raw = json.load(f)

                # Support three formats:
                # 1. Plain list of row dicts
                # 2. Dict with a "rows" / "entries" / "data" key
                # 3. Metadata-only dict (no phrase rows yet)
                if isinstance(raw, list):
                    self.data = raw
                    self.metadata = {}
                elif isinstance(raw, dict):
                    self.metadata = {k: v for k, v in raw.items()
                                     if not isinstance(v, list) or k in
                                     ("languages", "columns", "cleaning_rules_applied")}
                    for key in ("rows", "entries", "data", "phrases"):
                        candidate_rows = raw.get(key)
                        if not (
                            isinstance(candidate_rows, list)
                            and candidate_rows
                            and isinstance(candidate_rows[0], dict)
                        ):
                            continue
                        if any(
                            _normalize_column(str(column_name))
                            for column_name in candidate_rows[0].keys()
                        ):
                            self.data = candidate_rows
                            break

                if self.data:
                    self.data = self._normalize_rows(self.data)
                    json_loaded = True

            except Exception as exc:
                safe_print(f"[Translation] Error loading JSON dataset: {exc}")
        else:
            safe_print(f"[Translation] ERROR: Dataset file not found at {self.dataset_path}")

        # --- Attempt 2: Scan backend directory for CSV/Excel ---
        if not json_loaded:
            base = os.path.dirname(self.dataset_path)
            candidates = self._find_data_files(base)
            for candidate_path in candidates:
                ext = os.path.splitext(candidate_path)[1].lower()
                if ext == ".csv":
                    self.data = self._load_csv(candidate_path)
                elif ext in (".xlsx", ".xls"):
                    self.data = self._load_excel(candidate_path)
                if self.data:
                    self.data = self._normalize_rows(self.data)
                    break

        # --- Common outcome ---
        if not self.data:
            safe_print(
                "[Translation] Dataset file has no phrase rows yet. "
                "Add a 'rows' array to translation_data.json, or place a "
                "translation_data.csv / phrasebook.csv in the backend folder."
            )
            return

        self._build_all_indices()
        self.is_loaded = True
        safe_print(f"[Translation] Dataset loaded: {len(self.data)} entries")
        safe_print("[Translation] Sample translations (first 5):")
        for i, row in enumerate(self.data[:5]):
            en = row.get("english_source", "")
            tg = row.get("tagabawa_source", "")
            fi = row.get("filipino_source", "")
            safe_print(f"  [{i+1}] en={en!r}  tagabawa={tg!r}  filipino={fi!r}")

    # ------------------------------------------------------------------
    # File discovery
    # ------------------------------------------------------------------

    def _find_data_files(self, directory: str) -> List[str]:
        """Return ordered list of candidate CSV/Excel paths to try."""
        preferred_csv = ["translation_data.csv", "phrasebook.csv", "dataset.csv"]
        preferred_xlsx = ["translation_data.xlsx", "phrasebook.xlsx", "dataset.xlsx"]

        found = []
        for name in preferred_csv:
            p = os.path.join(directory, name)
            if os.path.exists(p):
                found.append(p)

        for p in _glob.glob(os.path.join(directory, "*.csv")):
            if p not in found:
                found.append(p)

        for name in preferred_xlsx:
            p = os.path.join(directory, name)
            if os.path.exists(p):
                found.append(p)

        for p in _glob.glob(os.path.join(directory, "*.xlsx")):
            if p not in found:
                found.append(p)

        return found

    # ------------------------------------------------------------------
    # CSV loader
    # ------------------------------------------------------------------

    def _load_csv(self, path: str) -> List[Dict]:
        """Load phrase rows from a CSV file. Returns list of row dicts or []."""
        rows = []
        safe_print(f"[Translation] Loading dataset from: {path}")
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = _csv.DictReader(f)
                if reader.fieldnames is None:
                    safe_print(f"[Translation] ERROR: CSV has no header row: {path}")
                    return []

                # Build column map: csv_header -> canonical_key
                col_map = {}
                for h in reader.fieldnames:
                    canonical = _normalize_column(h)
                    if canonical:
                        col_map[h] = canonical

                if not col_map:
                    safe_print(f"[Translation] ERROR: No recognized language columns in CSV: {path}")
                    safe_print(f"[Translation] Found headers: {list(reader.fieldnames)}")
                    safe_print("[Translation] Expected headers like: tagabawa, english, filipino, cebuano")
                    return []

                for raw_row in reader:
                    row = {}
                    for csv_col, canonical_key in col_map.items():
                        val = (raw_row.get(csv_col) or "").strip()
                        row[canonical_key] = val
                    if any(row.values()):
                        rows.append(row)

        except Exception as exc:
            safe_print(f"[Translation] ERROR loading CSV {path}: {exc}")
        return rows

    # ------------------------------------------------------------------
    # Row normalization
    # ------------------------------------------------------------------

    def _normalize_rows(self, rows: List[Dict]) -> List[Dict]:
        """Ensure every supported language is available as {lang}_source.

        The bundled JSON uses plain language keys such as `english` and
        `tagabawa`, while the lookup indices use canonical keys such as
        `english_source` and `tagabawa_source`. CSV/Excel loading already
        produces canonical keys, but normalizing all loaded rows keeps the
        JSON and CSV paths equivalent.
        """
        normalized_rows: List[Dict] = []
        for raw_row in rows:
            if not isinstance(raw_row, dict):
                continue

            row = dict(raw_row)
            for header, value in raw_row.items():
                canonical_key = _normalize_column(str(header))
                if not canonical_key:
                    continue
                current_value = str(row.get(canonical_key, "") or "").strip()
                if current_value:
                    continue
                row[canonical_key] = str(value or "").strip()

            if any(str(row.get(f"{lang}_source", "") or "").strip() for lang in SUPPORTED_LANGS):
                normalized_rows.append(row)

        return normalized_rows

    # ------------------------------------------------------------------
    # Excel loader
    # ------------------------------------------------------------------

    def _load_excel(self, path: str) -> List[Dict]:
        """Load phrase rows from an .xlsx file. Returns list of row dicts or []."""
        rows = []
        safe_print(f"[Translation] Loading dataset from: {path}")
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            raw_headers = [
                str(cell.value or "").strip()
                for cell in next(ws.iter_rows(max_row=1))
            ]

            col_map = {}  # column index -> canonical_key
            for idx, h in enumerate(raw_headers):
                canonical = _normalize_column(h)
                if canonical:
                    col_map[idx] = canonical

            if not col_map:
                safe_print(f"[Translation] ERROR: No recognized language columns in Excel: {path}")
                safe_print(f"[Translation] Found headers: {raw_headers}")
                safe_print("[Translation] Expected headers like: tagabawa, english, filipino, cebuano")
                wb.close()
                return []

            for raw_row in ws.iter_rows(min_row=2, values_only=True):
                row = {}
                for idx, canonical_key in col_map.items():
                    val = str(raw_row[idx] or "").strip() if idx < len(raw_row) else ""
                    row[canonical_key] = val
                if any(row.values()):
                    rows.append(row)

            wb.close()
        except ImportError:
            safe_print("[Translation] ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0")
        except Exception as exc:
            safe_print(f"[Translation] ERROR loading Excel {path}: {exc}")
        return rows

    # ------------------------------------------------------------------
    # Index building
    # ------------------------------------------------------------------

    def _build_all_indices(self):
        self._phrase_indices = {lang: {} for lang in SUPPORTED_LANGS}
        self._word_indices = {lang: {} for lang in SUPPORTED_LANGS}

        for row in self.data:
            for lang in SUPPORTED_LANGS:
                text = row.get(f"{lang}_source", "").strip()
                if not text:
                    continue

                # Phrase index
                for norm in _lookup_variants(text):
                    self._phrase_indices[lang].setdefault(norm, []).append(row)

                # Word index
                for word in re.findall(r"\b\w+\b", _normalize_lookup_text(text, strip_diacritics=True)):
                    self._word_indices[lang].setdefault(word, []).append(row)

    # ------------------------------------------------------------------
    # Translation
    # ------------------------------------------------------------------

    def translate_phrase(
        self,
        text: str,
        source_lang: str = "english",
        target_lang: str = "tagabawa",
    ) -> str:
        """
        Translate text using: exact match -> fuzzy match -> word-by-word.
        Returns original text if no translation is found.
        """
        return self.translate_phrase_with_metadata(
            text,
            source_lang=source_lang,
            target_lang=target_lang,
        )["translated"]

    def translate_phrase_with_metadata(
        self,
        text: str,
        source_lang: str = "english",
        target_lang: str = "tagabawa",
    ) -> Dict:
        """
        Translate text and return frontend/API metadata.

        Returns:
            {
              translated,
              method,
              cascade_stage,
              confidence
            }
        """
        original = text.strip()
        if not self.is_loaded or not text.strip():
            return _translation_result(UNKNOWN_FOR_REVIEW, "unknown_for_review", 0.0)

        # Normalize language aliases
        source_lang = _normalize_lang(source_lang)
        target_lang = _normalize_lang(target_lang)

        if source_lang == target_lang:
            return _translation_result(text, "identity", 1.0)

        target_key = f"{target_lang}_source"
        src_index = self._phrase_indices.get(source_lang, {})

        # 1. Exact match
        for norm in _lookup_variants(original):
            result = _first_result(src_index.get(norm, []), target_key)
            if result:
                return _translation_result(
                    _preserve_case(original, result),
                    "exact_phrase",
                    1.0,
                )

        # 2. Fuzzy match (rapidfuzz)
        fuzzy = self._fuzzy_match(
            _normalize_lookup_text(original, strip_diacritics=True),
            src_index,
            target_key,
        )
        result = fuzzy["translated"] if fuzzy else None
        if result:
            return _translation_result(
                _preserve_case(original, result),
                "fuzzy_phrase",
                fuzzy["score"] / 100.0,
            )

        # 3. Word-by-word fallback
        word_result = self._translate_words(original, source_lang, target_lang)
        if word_result != original:
            return _translation_result(word_result, "word_by_word", 0.55)
        return _translation_result(UNKNOWN_FOR_REVIEW, "unknown_for_review", 0.0)

    def _fuzzy_match(
        self,
        query: str,
        src_index: Dict[str, List[Dict]],
        target_key: str,
        threshold: int = 82,
    ) -> Optional[Dict]:
        if not src_index:
            return None
        try:
            from rapidfuzz import process as rfp, fuzz  # type: ignore
            match = rfp.extractOne(
                query, list(src_index.keys()), scorer=fuzz.ratio
            )
            if match and match[1] >= threshold:
                translated = _first_result(src_index[match[0]], target_key)
                if translated:
                    return {"translated": translated, "score": float(match[1])}
        except ImportError:
            pass
        return None

    def _translate_words(
        self, text: str, source_lang: str, target_lang: str
    ) -> str:
        if not text.strip():
            return text

        leading = len(text) - len(text.lstrip())
        trailing = len(text) - len(text.rstrip())
        content = text.strip()

        # Preserve list prefixes (1. / a. / * etc.)
        prefix = ""
        m = re.match(r"^(\d+[.)]\s*|\w[.)]\s*|[*\-]\s+)", content)
        if m:
            prefix = m.group(0)
            content = content[len(prefix):]

        target_key = f"{target_lang}_source"
        word_idx = self._word_indices.get(source_lang, {})

        parts = re.split(r"(\s+|[,.\-;:!?])", content)
        translated: List[str] = []

        for part in parts:
            if not part:
                continue
            if re.match(r"^[\s,.\-;:!?]+$", part):
                translated.append(part)
                continue

            clean = _normalize_lookup_text(part, strip_diacritics=True)
            t = _first_result(word_idx.get(clean, []), target_key)
            if t:
                word = t.strip().split()[0]
                translated.append(_preserve_case(part, word))
            else:
                translated.append(part)

        result = prefix + "".join(translated)
        return " " * leading + result + " " * trailing

    # ------------------------------------------------------------------
    # Quick helpers used by the API
    # ------------------------------------------------------------------

    def translate_quick(
        self, text: str, source_lang: str = "english", target_lang: str = "tagabawa"
    ) -> Dict:
        translated = self.translate_phrase(text, source_lang, target_lang)
        return {
            "original": text,
            "translated": translated,
            "source_language": source_lang,
            "target_language": target_lang,
            "dataset_loaded": self.is_loaded,
        }


# ------------------------------------------------------------------
# Internal helpers
# ------------------------------------------------------------------

def _normalize_lang(lang: str) -> str:
    normalized = _normalize_alias_key(lang)
    hyphenated = normalized.replace(" ", "-")
    underscored = normalized.replace(" ", "_")
    return (
        _LANG_ALIASES.get(normalized)
        or _LANG_ALIASES.get(hyphenated)
        or _LANG_ALIASES.get(underscored)
        or normalized
    )


def _first_result(rows: List[Dict], key: str) -> Optional[str]:
    for row in rows:
        val = row.get(key, "").strip()
        if val:
            return val
    return None


def _translation_result(translated: str, method: str, confidence: float) -> Dict:
    return {
        "translated": translated,
        "method": method,
        "cascade_stage": method,
        "confidence": round(max(0.0, min(1.0, float(confidence))), 4),
    }


def _preserve_case(original: str, translated: str) -> str:
    if not translated:
        return translated
    if original and original[0].isupper():
        return translated[0].upper() + translated[1:]
    return translated


# ------------------------------------------------------------------
# Singleton
# ------------------------------------------------------------------

_dataset: Optional[TranslationDataset] = None


def get_translation_dataset() -> TranslationDataset:
    global _dataset
    if _dataset is None:
        _dataset = TranslationDataset()
    return _dataset


def translate(text: str, target_lang: str = "tagabawa") -> str:
    return get_translation_dataset().translate_phrase(
        text, source_lang="english", target_lang=target_lang
    )
